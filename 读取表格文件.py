from openpyxl import load_workbook
import settings

wb = load_workbook(filename=settings.SHEET_PATH)
ws = wb[wb.sheetnames[0]]
user_settings = {}
# 迭代所有的行
for row in ws.rows:
    if "T" in row[0].value:
        tool_num = row[0].value
        s = row[1].value
        f = row[2].value
        model = row[3].value
        val1 = row[4].value
        val2 = row[5].value
        user_settings[tool_num] = {
            "feed": f,
            "rspeed": s,
            "model": model,
            "var1": val1,
            "var2": val2,
        }

print(user_settings)


